import psycopg2
from config import config
from psycopg2 import sql
import pandas
from openpyxl import load_workbook

DATABASE_URL = config.get_db_url()

db_name = config.get_db_name()
db_user = config.get_username()
db_password = config.get_password()


class DB():
    def __init__(self, name:str , user:str , password:str):
        self.connect = psycopg2.connect(f'dbname={name} user={user} password={password}')
        self.cur = self.connect.cursor()

    def create_table(self, name:str, xlsx_file_path:str, sheet:str = None) -> None:
        wb = load_workbook(xlsx_file_path, read_only=True)
        sheet = wb[sheet] if sheet else wb.active
        columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        wb.close()

        columns_def = ",\n".join([f'{col} TEXT' for col in columns])
        create_table_query = f'''CREATE TABLE IF NOT EXISTS {name} 
        ({columns_def})'''

        cur = self.cur
        cur.execute(create_table_query)
        self.connect.commit()
        return None
    
    def xlsx_to_posgresql(self, xlsx_file_path: str, table_name: str, sheet_name: str = None) -> bool:
        """Импортируем данные из эксель в таблицу SQL"""
        wb = load_workbook(xlsx_file_path, read_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active
        
        try:
            columns = [str(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            placeholders = ','.join(['%s'] * len(columns))
            columns_str = ','.join(f'"{col}"' for col in columns)
            
            insert_query = f'INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders})'.lower()
            
            cursor = self.cur
            
            batch = []
            for row in sheet.iter_rows(min_row=2):
                values = [cell.value for cell in row]
                batch.append(values)
                
                # Вставляем пачками по 100 строк
                if len(batch) >= 100:
                    cursor.executemany(insert_query, batch)
                    self.connect.commit()
                    batch = []
            
            # Вставляем оставшиеся строки
            if batch:
                cursor.executemany(insert_query, batch)
                self.connect.commit()
                
            print("Данные успешно импортированы")
            return True
            
        except Exception as e:
            self.connect.rollback()
            print(f"Ошибка при импорте данных: {e}")
            return False
            
        finally:
            if hasattr(self, 'connect') and self.connect:
                self.connect.close()
            wb.close()

        
db = DB(name=db_name, user=db_user, password=db_password)
